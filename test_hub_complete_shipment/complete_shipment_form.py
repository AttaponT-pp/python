import sys
import re
from os import path
import datetime, time

import excelrd as xlrd
from openpyxl import load_workbook
from PyQt5 import QtWidgets
from PyQt5.QtCore import QTimer
import Form

import ctypes
import FITS_Connect as FITSDLL

global BG_color
BG_color = {'red':'background-color: rgb(255, 0, 0);', 'green': 'background-color: rgb(0, 255, 0);', 'blue': 'background-color: rgb(0, 0, 255);',
            'yellow': 'background-color: rgb(255, 255, 0);', 'white': 'background-color: rgb(255, 255, 255);','no_fill': 'background-color: rgb();'}

class CompleteShipmentMainUi(QtWidgets.QMainWindow):

    def __init__(self):
        super(CompleteShipmentMainUi, self).__init__()
        self.form = Form.Ui_MainForm()
        self.form.setupUi(self)
        self.setFixedSize(380, 340)
        # Init timer
        self.flag = True
        self.timer = QTimer(self, interval=450)
        self.timer.timeout.connect(self.blink_en)
        self.timer.start()
        self.form.tabWidget.setCurrentIndex(0)
        self.fill_en()
        # connect events
        self.form.txten.returnPressed.connect(self.page_focus)
        self.form.tabWidget.currentChanged.connect(self.tab_on_change)
        self.form.btn_oba_browse.clicked.connect(self.select_oba_file)
        self.form.btn_rtv_browse.clicked.connect(self.select_rtv_file)
        self.form.txt_inv.textChanged.connect(self.inv_count)
        self.form.txt_etr.textChanged.connect(self.etr_count)
        self.form.txt_inv.returnPressed.connect(self.get_invoice)
        self.form.txt_etr.returnPressed.connect(self.get_etr)

    def tab_on_change(self):
        # get en
        en = self.form.txten.text()
        # get tab index
        tab_index = self.form.tabWidget.currentIndex()
        # get checkbox
        which_opn = self.check_opn_box()
        # EN validation
        if len(en) != 6:
            self.fill_en()
        else:
            # validate EN certified
            if tab_index == 0:
                if which_opn['opn1501'] and not is_en_cert(FITSDLL, '1501', en)['status']:
                    mbox(title='EN Certified Error', text='{} of Operation:{}'.format(is_en_cert(FITSDLL, '1501', en)['msg'], '1501'), style=0)
                    self.fill_en()
                    return
                if which_opn['opn702'] and not is_en_cert(FITSDLL, '702', en)['status']:
                    mbox(title='EN Certified Error', text='{} of Operation:{}'.format(is_en_cert(FITSDLL, '702', en)['msg'], '702'), style=0)
                    self.fill_en()
                    return
            else:
                if which_opn['opn1502'] and not is_en_cert(FITSDLL, '1502', en)['status']:
                    mbox(title='EN Certified Error', text='{} of Operation:{}'.format(is_en_cert(FITSDLL, '1502', en)['msg'], '1502'), style=0)
                    self.fill_en()
                    return
                if which_opn['opn1801'] and not is_en_cert(FITSDLL, '1801', en)['status']:
                    mbox(title='EN Certified Error', text='{} of Operation:{}'.format(is_en_cert(FITSDLL, '1801', en)['msg'], '1801'), style=0)
                    self.fill_en()
                    return
            # input EN is valid
            self.timer.stop()    
            self.form.txten.setStyleSheet(BG_color['white'])
            self.form.textEdit_5.setStyleSheet(BG_color['green'])            
            if tab_index == 0:
                print('Page 0')
                if self.form.oba_file_path.toPlainText() == '':
                    self.form.btn_oba_browse.setFocus()
                    self.form.lbl_app_status.setText('Browse your OBA Request File.')
                else:
                    self.form.txt_inv.setFocus()
                    self.form.txt_inv.setText('')
                    self.form.lbl_app_status.setText('Scan Invoice#')
            else:
                print('Page 1')
                if self.form.rtv_file_path.toPlainText() == '':
                    self.form.btn_rtv_browse.setFocus()
                    self.form.lbl_app_status.setText('Browse your Daily Shipment Request File.')
                else:
                    self.form.txt_etr.setFocus()
                    self.form.txt_etr.setText('')
                    self.form.lbl_app_status.setText('Scan ETR No.')

    def get_invoice(self):
        now = datetime.datetime.now()
        print('Today: {}',format(now))
        self.form.textEdit.setStyleSheet(BG_color['no_fill'])
        self.form.textEdit_2.setStyleSheet(BG_color['no_fill'])
        f_path = str(self.form.oba_file_path.toPlainText())
        filled_en = str(self.form.txten.text())
        f_data = {'rt': '', 'packing_lot': '', 'po_num': '', 'part_num': '', 'qty': '', 'inv': ''}

        # Check EN
        if filled_en == "" or not len(filled_en) == 6:
            mbox(u'Please Fill EN Again', u'Please Fill EN Again', 0)
            self.form.txt_etr.setText("")
            self.fill_en()
            return
        # Check file path
        if f_path == "":
            mbox(u'Not Found OBA Request File', u'Please Browse OBA Request before Enter ETR', 0)
            self.form.txt_inv.setText("")
            return
        # Check file exist
        if not path.isfile(f_path):
            mbox(u'Not Found OBA Request File', u'File Not Found! Please browse the file again', 0)
            self.form.txt_inv.setText("")
            self.form.oba_file_path.setText("")
            return

        inv_num = str(self.form.txt_inv.text().strip())

        # check length of Invoice#
        if len(inv_num) != 7:
            mbox('Wrong invoice number format', "Please re-scan invoice number", 0)
            self.form.txt_inv.setText('')
            return

        # check invoice# in OBA Request
        if not cross_check_inv(f_path, inv_num, f_data):
            mbox(u'Not Found Invoice number', u'Not Found Invoice# in OBA Request', 0)
            self.form.txt_inv.setText('')
            return

        if f_data['inv'] != '':
            tmp_inv = f_data['inv']
            print('Invoice no. {}'.format(tmp_inv))
            self.form.lbl_app_status.setText('Found Invoice# {}'.format(tmp_inv))
            # get check box operation
            which_opn = self.check_opn_box()
            if not which_opn['opn1501'] and not which_opn['opn702']:
                mbox('No FITS OPN select.', 'Not found FITS OPN selected.\n'
                     'Please select target OPN.', 0)
                return
            if which_opn['opn1501']:
                # Opn.1501 OBA
                # check packing_lot
                if f_data['packing_lot'] == 'None' or f_data['packing_lot'] is None:
                    print('Packing Number is empty.')
                    # get Packing Number from FITS
                    f_data['packing_lot'] = FITSDLL.find_packing_num(FITSDLL.fits_dll, f_data['rt'])
                    if f_data['packing_lot'] == False:
                        self.form.textEdit.setStyleSheet(BG_color['no_fill'])
                        mbox('FITSDLL Error', 'Cannot initial FITSDLL.', 0)
                        return
                    elif f_data['packing_lot'] == 'DTS':
                        self.form.textEdit.setStyleSheet(BG_color['no_fill'])
                        mbox('Dock To Stock (DTS) detected', 'The build type of RT: {} is {}.\n'
                             'No need to save data to operation 1501 OBA.'.format(f_data['rt'], f_data['packing_lot']), 0)
                        return
                    
                # prepare data_stream
                packing_no = f_data['packing_lot']
                oba_info = FITSDLL.prepare_oba_info(FITSDLL.fits_dll, packing_no)
                print(oba_info)

                # create data1501
                data1501 = filled_en + ',' + inv_num + ',' + f_data['packing_lot'] + ',' + oba_info
                print('Opn.1501 param: {}'.format(FITSDLL.opn1501_param))
                print('Data stream: {}'.format(data1501))
                # hand_check data and operation
                inv_last_opn = FITSDLL.get_last_opn(FITSDLL.fits_dll, inv_num)
                print(inv_last_opn)
                route_check = FITSDLL.valid_inv('1501', inv_num)
                print('Hand-Check: {}'.format(route_check['status']))
                if route_check['status']:
                    # save FITS data
                    print('Recording data...')
                    if FITSDLL.record2fit(FITSDLL.fits_dll, '1501', FITSDLL.opn1501_param, data1501):
                        print('Save data opn.1501 OBA completed.')
                        self.form.txt_inv.setFocus()
                        self.form.txt_inv.setText('')
                        self.form.textEdit.setStyleSheet(BG_color['green'])
                        self.form.lbl_app_status.setText('Invoice no.{} is saved successful.'.format(inv_num))
                    else:
                        print('Cannot save data opn.1501 OBA.')
                        self.form.txt_inv.setFocus()
                        self.form.txt_inv.setText('')
                        self.form.textEdit.setStyleSheet(BG_color['red'])
                        self.form.lbl_app_status.setText('FITS Error: Cannot save opn.1501')
                        return
                else:
                    self.form.lbl_app_status.setText('FITSDLL Error: {}'.format(route_check["msg"]))
                    self.form.textEdit.setStyleSheet(BG_color['red'])
                    mbox('FITSDLL Error', route_check["msg"], 0)
                    return

            if which_opn['opn702']:
                # Check DTS
                if f_data['packing_lot'] is None:
                    print('Packing Number is empty.')
                    # get Packing Number from FITS
                    f_data['packing_lot'] = FITSDLL.find_packing_num(FITSDLL.fits_dll, f_data['rt'])
                    if f_data['packing_lot'] == False:
                        self.form.textEdit.setStyleSheet(BG_color['no_fill'])
                        self.form.textEdit_2.setStyleSheet(BG_color['no_fill'])
                        mbox('FITSDLL Error', 'Cannot initial FITSDLL.', 0)
                        return
                    elif f_data['packing_lot'] == 'DTS':
                        self.form.textEdit.setStyleSheet(BG_color['no_fill'])
                        self.form.textEdit_2.setStyleSheet(BG_color['no_fill'])
                        mbox('Dock To Stock (DTS) detected', 'The build type of RT: {} is {}.\nNo need to save data to operation 702 ShipmentSN.'.format(f_data['rt'], f_data['packing_lot']), 0)
                        return
                # Opn.702 Shipment SN
                sn_list = FITSDLL.get_sn_list(FITSDLL.fits_dll, f_data['rt'])
                print(sn_list)
                for sn in sn_list.split(','):
                    # print('Serial No.:{}'.format(sn))
                    print('Validate route...')
                    route_check = FITSDLL.valid_inv('702', sn)
                    print(route_check['status'])
                    if route_check['status']:
                        # get last operation
                        print('Get last operation...')
                        last_opn = FITSDLL.get_last_opn(FITSDLL.fits_dll, sn)
                        print('Last operation of SN: {} is {}'.format(sn, last_opn))
                        if last_opn == '601_B':
                            print('Ready to save...')
                            # prepare data_stream
                            data702 = filled_en + ',' + inv_num + ',' + sn + ',' + f_data['packing_lot'] + ',' + str(f_data['qty'])
                            print('Data stream: {}'.format(data702))
                            print('Opn.702 param: {}'.format(FITSDLL.opn702_param))
                            print('Record data...')
                            if FITSDLL.record2fit(FITSDLL.fits_dll, '702', FITSDLL.opn702_param, data702):
                                print('Save SN: {} successful.'.format(sn))
                                self.form.txt_inv.setFocus()
                                self.form.txt_inv.setText('')
                                self.form.textEdit_2.setStyleSheet(BG_color['green'])
                                self.form.lbl_app_status.setText('Invoice no.{} : {} successful.'.format(inv_num, sn))
                            else:
                                print('Save SN: {} error.'.format(sn))
                                self.form.txt_inv.setFocus()
                                self.form.txt_inv.setText('')
                                self.form.textEdit.setStyleSheet(BG_color['no_fill'])
                                self.form.textEdit_2.setStyleSheet(BG_color['red'])
                                self.form.lbl_app_status.setText('FITS Error: Cannot save opn.702')
                                return
                        else:
                            self.form.textEdit_2.setStyleSheet(BG_color['red'])
                            self.form.lbl_app_status.setText('This SN: {} does not pack yet'.format(sn))
                            print('This SN: {} does not pack yet'.format(sn))
                    else:
                        print('SN:{} route check error.'.format(sn))
                        self.form.lbl_app_status.setText('SN: {} route check error.'.format(sn))
                        self.form.textEdit_2.setStyleSheet(BG_color['red'])
        else:
            mbox('Not found Invoice# {}'.format(inv_num), 'Not found Invoice# {} in OBA Summary File'.format(inv_num), 0)
            return

    def get_etr(self):
        now = datetime.datetime.now()
        print('Today: {}'.format(now))
        self.form.textEdit_3.setStyleSheet(BG_color['no_fill'])
        self.form.textEdit_4.setStyleSheet(BG_color['no_fill'])
        f_path = str(self.form.rtv_file_path.toPlainText())
        filled_en = str(self.form.txten.text())
        f_data = {'inv': '', 'rt': '', 'qty': ''}

        # Check EN
        if filled_en == "" or not len(filled_en) == 6:
            mbox('Please Fill EN Again', 'Please Fill EN Again', 0)
            self.form.txt_etr.setText("")
            self.fill_en()
            return

        # Check file path
        if f_path == "":
            mbox('Not Found File Daily Shipment Request', 'Please Browse Daily Shipment Request before Enter ETR', 0)
            self.form.txt_etr.setText("")
            return

        # Check file exist
        if not path.isfile(f_path):
            mbox('Not Found File Daily Shipment Request', 'File Not Found! Please browse the file again', 0)
            self.form.txt_etr.setText("")
            self.form.rtv_file_path.setText("")
            return

        etr = str(self.form.txt_etr.text())

        # Check prefix of ETR
        if not etr[0].upper() == "C" and not etr[0].upper() == "R":
            mbox('ETR wrong format', 'Please Enter ETR in correct format', 0)
            self.form.txt_etr.setText("")
            return

        # Check across this year
        # if not etr[1:5] == str(now.year):
        # Mbox(u'ETR wrong format', u'Please Enter ETR in correct format', 0)
        # self.line_etr.setText("")
        # return

        # Check length of ETR
        if not len(etr) == 9:
            mbox('ETR wrong format', 'Please Enter ETR in correct format', 0)
            self.form.txt_etr.setText("")
            return

        # Check ETR across file
        if not cross_check_etr(f_path, etr, f_data):
            mbox(u'Not Found ETR', 'Not Found ETR in File Daily Shipment Request', 0)
            self.form.txt_etr.setText("")
            return

        # Print Invoice on Label
        # if f_data['inv'] != "None" and f_data['rt'] != "None" and f_data['qty'] != "None":
        if f_data['inv'] != "None" or f_data['inv'] is not None:
            tmp_inv = f_data['inv']
            # Get RTV Shipment Blocking
            data1502 = str(FITSDLL.prepare_etr_info(FITSDLL.fits_dll, etr))
            block_rtv_status = data1502.split(',')[len(data1502.split(','))-1]
            if block_rtv_status == "YES":
                mbox("Warning !!!", "This ETR Number:" + etr + ". have been blocked in Opn.924 RTV Shipment Blocking"
                                                               "Please inform case owner for unblock.", 0)
                self.form.lbl_app_status.setText("ETR Number:" + etr + " have been blocked RTV Shipment")
                self.form.txt_etr.setFocus()
                self.form.txt_etr.setText("")
                return

            self.form.lbl_app_status.setText('Found Invoice {}'.format(tmp_inv))
            which_opn = self.check_opn_box()

            if not which_opn["opn1502"] and not which_opn["opn1801"]:
                mbox('No FITS OPN Selected', 'Please Select FITS Operation to record.', 0)
                return

            if which_opn["opn1502"]:
                # Create data_str input stream
                data_str = filled_en + ',' + tmp_inv + ',' + etr + ',' + block_rtv_status
                fit_status = FITSDLL.valid_inv('1502', tmp_inv)
                if fit_status["status"]:
                    # Add Parameter 'RTV Shipment Blocking'
                    param1502_str = 'OPERATOR,Invoice No,ETR Number,RTV Shipment Blocking'
                    print(param1502_str)
                    print(data_str)
                    # if record2fit('1502', param1502_param, data_str):
                    if FITSDLL.record2fit(FITSDLL.fits_dll, '1502', param1502_str, data_str):
                        print('ETR No.{} is save with Invoice No.{} successful.'.format(etr, tmp_inv))
                        self.form.lbl_app_status.setText('FITS1502 Saved for {}'.format(etr))
                        self.form.txt_etr.setText("")
                        self.form.txt_etr.setFocus()
                        self.form.textEdit_3.setStyleSheet(BG_color['green'])
                    else:
                        mbox('FITSDLL Error', 'Cannot save data to FITS', 0)
                        self.form.textEdit_3.setStyleSheet(BG_color['red'])
                        return
                else:
                    mbox('FITSDLL Error', fit_status["msg"], 0)
                    self.form.textEdit_3.setStyleSheet(BG_color['red'])
                    return

            if which_opn["opn1801"]:
                time.sleep(1)
                FITSDLL.init('1801')
                # Create data_str input stream
                data_str = filled_en + ',' + tmp_inv + ',' + etr
                data1303 = FITSDLL.get_necessory_data(FITSDLL.fits_dll, '1303', etr, 'Serial No,Fail Qty')
                print(data1303)
                if FITSDLL.init('1801') == 'True':
                    param1801_str = 'OPERATOR,Invoice No,ETR Number,Serial No,ETR Lot Qty'
                    if len(data1303) < 3:
                        ship_data_str = data_str + ',' + data1303[0] + ',' + data1303[1]
                        print(ship_data_str)

                        if FITSDLL.record2fit(FITSDLL.fits_dll, '1801', param1801_str, ship_data_str) == 'True':
                            self.form.lbl_app_status.setText('FITS1801 Saved for {}'.format(etr))
                            self.form.txt_etr.setText("")
                            self.form.txt_etr.setFocus()
                            self.form.textEdit_4.setStyleSheet(BG_color['green'])
                        else:
                            mbox('FITSDLL Error', 'Cannot save data to FITS', 0)
                            self.form.textEdit_4.setStyleSheet(BG_color['red'])
                            return
                    else:
                        for sn in data1303:
                            ship_data_str = ""
                            if len(sn) == 11:
                                ship_data_str = data_str + ',' + sn + ',' + data1303[len(data1303) - 1]
                                print(ship_data_str)

                                if FITSDLL.record2fit(FITSDLL.fits_dll, '1801', param1801_str, ship_data_str):
                                    self.form.lbl_app_status.setText('FITS1801 Saved for {}'.format(etr))
                                    self.form.txt_etr.setText("")
                                    self.form.txt_etr.setFocus()
                                    self.form.textEdit_4.setStyleSheet(BG_color['green'])
                                else:
                                    mbox('FITSDLL Error', 'Cannot save data to FITS', 0)
                                    self.form.textEdit_4.setStyleSheet(BG_color['red'])
                                    return

                else:
                    mbox('FITSDLL Error', 'Unable to init FITs DB', 0)
                    self.form.textEdit_4.setStyleSheet(BG_color['red'])
                    return

        else:
            mbox('Not Found Invoice/RT/Qty', 'Not Found Invoice/RT/Qty for ETR: {}'.format(etr), 0)
            return

            # 2nd get neccessory data from rt
            # param = 'Part No.,Supplier Name,Build Type,PO No.'
            # param1502_out = param.split(',')
            # print tmp_inv+',' +tmp_rt+','+ tmp_qty
            # data1502 = get_necessory_data('101',tmp_rt,param)
            # Valid output
            # for i in range(len(data1502)):

            # print tmp_param[i] + '=' + q_data[i] + '\n'
            # if data1502[i] == "-":
            # Mbox(u'No Data Found!', u'There is no data for {} from FIT.'.format(param1502_out[i]), 0)
            # self.CleanupCtrs()
            # return
            
    def fill_en(self):
        self.timer.start()
        self.form.textEdit_5.setStyleSheet(BG_color['no_fill'])
        self.form.lbl_app_status.setText('Please fill your EN')
        self.form.txten.setText('')
        self.form.txten.setFocus()
        return
    
    def blink_en(self):
        if self.flag:
            # yellow
            self.form.txten.setStyleSheet(BG_color['yellow'])
        else:
            # grey
            self.form.txten.setStyleSheet(BG_color['no_fill'])
        self.flag = not self.flag
        
    def page_focus(self):
        en = self.form.txten.text()
        tab_index = self.form.tabWidget.currentIndex()
        which_opn = self.check_opn_box()
        oba_path = str(self.form.oba_file_path.toPlainText())
        rtv_path = str(self.form.rtv_file_path.toPlainText())
        # validate en length
        if len(en) != 6:
            mbox("EN Validation", "Your EN is not valid.\nPlease try again.", 0)
            self.fill_en()        
        else:
            # validate EN certified
            if tab_index == 0:
                if which_opn['opn1501'] and not is_en_cert(FITSDLL, '1501', en)['status']:
                    mbox(title='EN Certified Error', text='{} of Operation:{}'.format(is_en_cert(FITSDLL, '1501', en)['msg'], '1501'), style=0)
                    self.fill_en()
                    return
                if which_opn['opn702'] and not is_en_cert(FITSDLL, '702', en)['status']:
                    mbox(title='EN Certified Error', text='{} of Operation:{}'.format(is_en_cert(FITSDLL, '702', en)['msg'], '702'), style=0)
                    self.fill_en()
                    return
            else:
                if which_opn['opn1502'] and not is_en_cert(FITSDLL, '1502', en)['status']:
                    mbox(title='EN Certified Error', text='{} of Operation:{}'.format(is_en_cert(FITSDLL, '1502', en)['msg'], '1502'), style=0)
                    self.fill_en()
                    return
                if which_opn['opn1801'] and not is_en_cert(FITSDLL, '1801', en)['status']:
                    mbox(title='EN Certified Error', text='{} of Operation:{}'.format(is_en_cert(FITSDLL, '1801', en)['msg'], '1801'), style=0)
                    self.fill_en()
                    return
            # input EN is valid
            self.timer.stop()
            self.form.txten.setStyleSheet(BG_color['white'])
            self.form.textEdit_5.setStyleSheet(BG_color['green'])
            if tab_index == 0:
                print('Page 0')
                if oba_path != '':
                    self.form.txt_inv.setFocus()
                    self.form.lbl_app_status.setText('Scan Invoice#')
                    return
                else:
                    self.form.btn_oba_browse.setFocus()
                    self.form.lbl_app_status.setText('Browse your OBA Request File.')
                    return
            else:
                print('Page 1')
                if rtv_path != '':
                    self.form.txt_etr.setFocus()
                    self.form.lbl_app_status.setText('Scan ETR No.')
                    return
                else:
                    self.form.btn_rtv_browse.setFocus()
                    self.form.lbl_app_status.setText('Browse your Daily Shipment Request File.')
                    return

    def select_oba_file(self):
        dlg = QtWidgets.QFileDialog.getOpenFileName()
        print(dlg)
        if dlg[0] == '':
            mbox('OBA Request File Selection', 'No OBA Request file select', 0)
            self.form.oba_file_path.setText(dlg[0])
        else:
            print('Selected file: {}'.format(dlg))
            self.form.oba_file_path.setText(dlg[0])
            self.form.txt_inv.setFocus()
            self.form.txt_inv.setText('')
            self.form.lbl_app_status.setText('Scan Invoice#')
            return

    def select_rtv_file(self):
        dlg = QtWidgets.QFileDialog.getOpenFileName()
        print(dlg)
        if dlg[0] == '':
            mbox('Daily Shipment Request File Selection', 'No Daily Shipment Request file select', 0)
            self.form.rtv_file_path.setText(dlg[0])
        else:
            print('Selected file: {}'.format(dlg))
            self.form.rtv_file_path.setText(dlg[0])
            self.form.txt_etr.setFocus()
            self.form.txt_etr.setText('')
            self.form.lbl_app_status.setText('Scan ETR No.')
            return
    
    def etr_count(self):
        self.form.lbl_etr_count.setText("({})".format(len(self.form.txt_etr.text())))

    def inv_count(self):
        self.form.lbl_inv_count.setText("({})".format(len(self.form.txt_inv.text())))

    def check_opn_box(self):
        return {"opn1501": self.form.checkBox1501.isChecked(), "opn702": self.form.checkBox702.isChecked(),
                "opn1502": self.form.checkBox1502.isChecked(), "opn1801": self.form.checkBox1801.isChecked()}


def mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)

def is_en_cert(fitsdll, opn, en):
    return FITSDLL.en_validate(opn=opn, en=en)

# using openpyxl
def cross_check_inv(xlsx_path, inv, f_data):
    print('Cross Check Invoice number# {}'.format(inv))
    oba_wb = load_workbook(xlsx_path)
    output = oba_wb.active
    print(output.title)
    for row in range(5, output.max_row):
        # get invoice#
        if output.cell(row=row, column=10).value is None:
            break
        tmp_inv = output.cell(row=row, column=10).value
        print(tmp_inv)
        if re.search(inv, tmp_inv, re.IGNORECASE):
            f_data['rt'] = output.cell(row=row, column=2).value
            f_data['packing_lot'] = output.cell(row=row, column=3).value
            f_data['po_num'] = output.cell(row=row, column=4).value
            f_data['part_num'] = output.cell(row=row, column=5).value
            f_data['qty'] = output.cell(row=row, column=6).value
            inv_num = "{}".format(output.cell(row=row, column=10).value)
            f_data['inv'] = inv_num.split('\n')[0]
            print(f_data['inv'])
            return True
    return False

def cross_check_etr(xls_path, etr, f_data):
    print('Cross Check ETR: {}'.format(etr))
    # check input excel file, using openpyxl
    if xls_path.endswith('.xlsx'):
        rtv_wb = load_workbook(xls_path)
        shipment = rtv_wb.active
        print(shipment.title)
        for row in range(5, shipment.max_row + 1):
            if shipment.cell(row=row, column=23).value is None:
                break
            # get ETR#
            tmp_etr = shipment.cell(row=row, column=23).value
            print(tmp_etr)
            if re.search(etr, tmp_etr, re.IGNORECASE):
                f_data['inv'] = str(shipment.cell(row=row, column=22).value)
                return True
        return False
    # using excelrd as xlrd
    elif xls_path.endswith('.xls'):
        rtv_workbook = xlrd.open_workbook(xls_path)
        print(rtv_workbook.get_sheets)
        shipment = rtv_workbook.sheet_by_name('Shipment')
        print(shipment.name)
        for row in range(0, shipment.nrows):
            hold = shipment.cell_value(rowx=row, colx=22)
            print(hold)
            if re.search(etr, hold, re.IGNORECASE):
                inv_num = "{}" .format(shipment.cell_value(row, 21))
                if "." in inv_num:
                    f_data['inv'] = inv_num.replace(".", "")
                elif " " in inv_num:
                    f_data['inv'] = inv_num.replace(" ", "")
                print(f_data['inv'])
                return True
        return False
    else:
        print('Wrong type of excel file selected.')
        return False
    

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    main_app = CompleteShipmentMainUi()
    main_app.show()
    sys.exit(app.exec_())